"""
nestube/nestube/nestube/cuts_export.py
Export cut list to an Excel (.xlsx) workbook.
"""
from __future__ import annotations

from typing import List

from .models import Corte


def export_cuts_to_excel(filepath: str, cortes: List[Corte]) -> None:
    """Write the cut list to an xlsx file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cut List"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Description",
        "Length (mm)",
        "Qty",
        "Bevel L",
        "Bevel L dir",
        "Bevel L angle (°)",
        "Bevel R",
        "Bevel R dir",
        "Bevel R angle (°)",
        "Total length (mm)",
    ]
    col_widths = [28, 14, 6, 9, 12, 18, 9, 12, 18, 16]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border
        ws.column_dimensions[chr(64 + col_idx)].width = width

    ws.row_dimensions[1].height = 20

    for row_idx, c in enumerate(cortes, 2):
        total = c.largo * c.cantidad
        row_data = [
            c.descripcion,
            round(c.largo, 2),
            c.cantidad,
            "Yes" if c.inglete1 else "No",
            c.inglete1_dir or "up",
            round(c.inglete1_deg or 0, 1),
            "Yes" if c.inglete2 else "No",
            c.inglete2_dir or "up",
            round(c.inglete2_deg or 0, 1),
            round(total, 2),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = cell_border
            if col_idx in (2, 6, 9, 10):
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in (3,):
                cell.alignment = Alignment(horizontal="center")

    wb.save(filepath)
