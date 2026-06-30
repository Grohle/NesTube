"""
nestube/nestube/nestube/excel_import.py
Excel import/export template for cut lists.
"""
from __future__ import annotations

from typing import List

from .models import Corte


TEMPLATE_COLUMNS = [
    "Name",
    "Length (mm)",
    "Quantity",
    "Bevel 1",
    "Bevel 1 Direction",
    "Bevel 1 Angle",
    "Bevel 2",
    "Bevel 2 Direction",
    "Bevel 2 Angle",
]

_INSTRUCTIONS_EN = (
    "HOW TO FILL IN THIS TEMPLATE\n"
    "\n"
    "• Name: short description of the cut piece.\n"
    "• Length (mm): piece length in millimetres.\n"
    "• Quantity: integer number of pieces.\n"
    "• Bevel 1 / Bevel 2: 'Yes' to enable, 'No' to disable.\n"
    "• Bevel Direction: 'up' or 'down' only.\n"
    "• Bevel Angle: degrees, e.g. 45 (range 5–85).\n"
    "Leave unused bevel columns blank or 'No'."
)

_INSTRUCTIONS_ES = (
    "CÓMO RELLENAR ESTA PLANTILLA\n"
    "\n"
    "• Name: descripción breve de la pieza cortada.\n"
    "• Length (mm): longitud de la pieza en milímetros.\n"
    "• Quantity: número entero de piezas.\n"
    "• Bevel 1 / Bevel 2: 'Yes' para activar, 'No' para desactivar.\n"
    "• Bevel Direction: solo 'up' o 'down'.\n"
    "• Bevel Angle: grados, p.ej. 45 (rango 5–85).\n"
    "Deja en blanco o 'No' los campos de bisel no usados."
)


def save_template(filepath: str) -> None:
    """Save an Excel template with headers, bevel validation dropdowns, and instructions."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cut Import Template"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_text = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_text
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for col_idx in range(1, len(TEMPLATE_COLUMNS) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 16

    # DataValidation: bevel enable columns (cols 4 and 7 = D and G).
    # sqref ranges MUST be SPACE-separated (OOXML); a comma raises
    # "TypeError: expected ...MultiCellRange" in modern openpyxl.
    yes_no_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    yes_no_dv.sqref = "D2:D1000 G2:G1000"
    ws.add_data_validation(yes_no_dv)

    # DataValidation: bevel direction columns (cols 5 and 8 = E and H)
    dir_dv = DataValidation(type="list", formula1='"up,down"', allow_blank=True)
    dir_dv.sqref = "E2:E1000 H2:H1000"
    ws.add_data_validation(dir_dv)

    # Instructions box in K2:M20 (outside the data table)
    thin = Side(style="thin")
    box_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    en_cell = ws.cell(row=2, column=11, value=_INSTRUCTIONS_EN)
    en_cell.alignment = Alignment(wrap_text=True, vertical="top")
    en_cell.font = Font(size=9)
    en_cell.border = box_border
    ws.column_dimensions["K"].width = 42
    ws.row_dimensions[2].height = 110

    es_cell = ws.cell(row=12, column=11, value=_INSTRUCTIONS_ES)
    es_cell.alignment = Alignment(wrap_text=True, vertical="top")
    es_cell.font = Font(size=9)
    es_cell.border = box_border
    ws.row_dimensions[12].height = 110

    wb.save(filepath)


def import_cuts_from_excel(filepath: str) -> List[Corte]:
    """Import cuts from an Excel file. Returns list of Corte objects."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except TypeError:
        # Some workbooks (e.g. re-saved by Excel) carry data validations openpyxl
        # can't parse in normal mode, raising
        # "TypeError: expected <class '...MultiCellRange'>". The read-only
        # streaming reader skips data validations, so retry with it.
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active

    cuts: List[Corte] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue

        name = str(row[0] or "").strip()
        if not name:
            continue

        try:
            length = float(row[1])
            qty = int(row[2])
        except (TypeError, ValueError):
            continue

        if length <= 0 or qty <= 0:
            continue

        bevel1 = _parse_bool(row[3] if len(row) > 3 else None)
        bevel1_dir = _parse_direction(row[4] if len(row) > 4 else None)
        bevel1_deg = _parse_angle(row[5] if len(row) > 5 else None)
        bevel2 = _parse_bool(row[6] if len(row) > 6 else None)
        bevel2_dir = _parse_direction(row[7] if len(row) > 7 else None)
        bevel2_deg = _parse_angle(row[8] if len(row) > 8 else None)

        cuts.append(Corte(
            descripcion=name,
            largo=length,
            cantidad=qty,
            inglete1=bevel1,
            inglete2=bevel2,
            inglete1_dir=bevel1_dir,
            inglete2_dir=bevel2_dir,
            inglete1_deg=bevel1_deg,
            inglete2_deg=bevel2_deg,
        ))

    try:
        wb.close()
    except Exception:
        pass
    return cuts


def _parse_bool(value) -> bool:
    if value is None:
        return False
    s = str(value).strip().lower()
    return s in ("yes", "si", "sí", "oui", "是", "ja", "1", "true", "x")


def _parse_direction(value) -> str:
    if value is None:
        return "up"
    s = str(value).strip().lower()
    if s in ("down", "abajo", "bas", "下", "unten"):
        return "down"
    return "up"


def _parse_angle(value) -> float:
    if value is None:
        return 45.0
    try:
        ang = float(value)
        return max(5.0, min(85.0, ang))
    except (TypeError, ValueError):
        return 45.0
